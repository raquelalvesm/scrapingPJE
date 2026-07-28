import os, re, time, unicodedata
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException, TimeoutException, StaleElementReferenceException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
import pyperclip

CAMINHO_EXCEL = r'C:\Users\pi100348\Documents\scrapping_pje\Pasta1.xlsx'
CAMINHO_SAIDA = r'C:\Users\pi100348\Documents\scrapping_pje\resultado.xlsx'
CAMINHO_PERFIL = r'C:\Users\pi100348\Documents\scrapping_pje\perfil_do_chrome'
NOME_PERFIL = 'Pagamento_Pericia'
COLUNAS_SAIDA = ['nr_processo', 'polo_passivo', 'profissao', 'nome', 'data_nomeacao', 'data_servico', 'valor']
MESES = {'janeiro': '01', 'fevereiro': '02', 'marco': '03', 'abril': '04', 'maio': '05', 'junho': '06', 'julho': '07', 'agosto': '08', 'setembro': '09', 'outubro': '10', 'novembro': '11', 'dezembro': '12'}


def log(msg):
    print(f'[{datetime.now().strftime("%H:%M:%S")}] {msg}')


def remover_acentos(texto):
    return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')


def ler_processos_excel(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    cabecalhos = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if 'nr_processo' not in cabecalhos:
        raise ValueError(f'Coluna nr_processo nao encontrada. Colunas: {cabecalhos}')
    idx = cabecalhos.index('nr_processo') + 1
    valores = []
    for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
        if row[0].value is not None:
            valores.append(str(row[0].value).strip())
    wb.close()
    return valores


def salvar_resultado(dados_lista, caminho_saida):
    log(f'Salvando {len(dados_lista)} resultados...')
    if os.path.exists(caminho_saida):
        wb = openpyxl.load_workbook(caminho_saida)
        ws = wb.active
        ultima_linha = ws.max_row
        for row in range(ws.max_row, 1, -1):
            if ws.cell(row=row, column=1).value is not None:
                ultima_linha = row
                break
        else:
            ultima_linha = 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Resultado'
        fonte_negrito = Font(bold=True)
        for col_idx, coluna in enumerate(COLUNAS_SAIDA, 1):
            cell = ws.cell(row=1, column=col_idx, value=coluna)
            cell.font = fonte_negrito
            cell.alignment = Alignment(horizontal='center')
        ultima_linha = 1

    existentes = {}
    for row in range(2, ultima_linha + 1):
        valor = ws.cell(row=row, column=1).value
        if valor:
            existentes[str(valor).strip()] = row

    for dados in dados_lista:
        nr = dados.get('nr_processo', '').strip()
        if nr in existentes:
            linha = existentes[nr]
            for col_idx, coluna in enumerate(COLUNAS_SAIDA, 1):
                ws.cell(row=linha, column=col_idx, value=dados.get(coluna, ''))
            log(f'  Atualizado: {nr} (linha {linha})')
        else:
            ultima_linha += 1
            for col_idx, coluna in enumerate(COLUNAS_SAIDA, 1):
                ws.cell(row=ultima_linha, column=col_idx, value=dados.get(coluna, ''))
            log(f'  Inserido: {nr} (linha {ultima_linha})')

    try:
        wb.save(caminho_saida)
        wb.close()
        log(f'Arquivo salvo em {caminho_saida}')
        linhas = sum(1 for _ in openpyxl.load_workbook(caminho_saida, read_only=True).active.iter_rows(min_row=2)) + 1
        log(f'Total de linhas no arquivo: {linhas}')
    except PermissionError:
        log('ERRO: Nao foi possivel salvar o arquivo. Feche o Excel e execute novamente.')
        wb.close()


def extrair_dados_documento(texto):
    dados = {}
    match_nr = re.search(r'PROCESSO[:\s]+(\d{7}-\d{2}\.\d{4}\.\d{2}\.\d{4})', texto)
    if match_nr:
        dados['nr_processo'] = match_nr.group(1).strip()
    match_reu = re.search(r'REU:\s*(.+)', texto)
    if match_reu:
        sigla = match_reu.group(1).strip()
        if ' - ' in sigla:
            sigla = sigla.rsplit(' - ', 1)[-1]
        dados['polo_passivo'] = sigla
    match_data_servico = re.search(r'para o dia (\d{2}/\d{2}/\d{4})', texto)
    if match_data_servico:
        dados['data_servico'] = match_data_servico.group(1)
    match_valor = re.search(r'R\$\s*([\d.,]+)', texto)
    if match_valor:
        dados['valor'] = match_valor.group(1).strip()
    def _fmt_profissao(raw):
        raw = re.sub(r'\([aA]\)', '', raw).strip()
        return ' '.join(w.capitalize() for w in raw.split())

    match_perito = re.search(r'(?<!\w)o\(a\)\s+(.+?)\s*Dr[\(a\)]+\.?\s+([A-Z][A-Z\s]+?)\s*\(.+?\)', texto)
    if match_perito:
        dados['profissao'] = _fmt_profissao(match_perito.group(1))
        dados['nome'] = match_perito.group(2).strip()
    else:
        match_perito = re.search(r'Dr[\(a\)]+\.?\s+([A-Z][A-Z\s]+?)\s*\((.+?)\)', texto)
        if match_perito:
            dados['nome'] = match_perito.group(1).strip()
            dados['profissao'] = match_perito.group(2).strip()
    match_data_nomeacao = re.search(r'(\d{1,2})\s+de\s+(janeiro|fevereiro|mar[çc]o|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro)\s+de\s+(\d{4})', texto, re.IGNORECASE)
    if match_data_nomeacao:
        dia = match_data_nomeacao.group(1).zfill(2)
        mes_nome = remover_acentos(match_data_nomeacao.group(2).lower())
        mes = MESES.get(mes_nome, '??')
        ano = match_data_nomeacao.group(3)
        dados['data_nomeacao'] = f'{dia}/{mes}/{ano}'
    return dados


def entrar_ngframe(driver):
    driver.switch_to.default_content()
    iframe = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#ngFrame')))
    driver.switch_to.frame(iframe)
    time.sleep(2)


def main():
    processos = ler_processos_excel(CAMINHO_EXCEL)
    log(f'Total de processos: {len(processos)}')
    for i, p in enumerate(processos, 1):
        log(f'  {i:3d} - {p}')

    perfil = os.path.join(CAMINHO_PERFIL, NOME_PERFIL)
    os.makedirs(perfil, exist_ok=True)
    options = webdriver.ChromeOptions()
    options.add_argument(f'--user-data-dir={perfil}')
    options.add_experimental_option('prefs', {'download.default_directory': os.getcwd(), 'download.prompt_for_download': False, 'download.directory_upgrade': True, 'profile.default_content_settings.popups': 0})
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(10)
    driver.get('https://pje1g.trf1.jus.br/pje/')
    log('Chrome aberto. Faca login no PJe manualmente.')
    input('Pressione ENTER apos estar no painel de tarefas do PJe... ')
    log('Login PJe confirmado.')

    try:
        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#ngFrame')))
        log('Painel de tarefas do PJe encontrado.')
    except TimeoutException:
        raise Exception('Nao foi possivel encontrar o painel de tarefas.')

    resultados = []
    for i, nr_processo in enumerate(processos, 1):
        log(f'  === Processo {i}/{len(processos)}: {nr_processo} ===')
        dados = {'nr_processo': nr_processo}

        try:
            driver.switch_to.window(driver.window_handles[0])
            entrar_ngframe(driver)
            menu = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#liConsultaProcessual a')))
            menu.click()
            time.sleep(2)
            pyperclip.copy(nr_processo)
            ActionChains(driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
            time.sleep(3)
            frame_consulta = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'frameConsultaProcessual')))
            driver.switch_to.frame(frame_consulta)
            time.sleep(1)
            botao = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#fPP\\:searchProcessos')))
            botao.click()
            time.sleep(3)
            log('Busca realizada com sucesso.')
            link_processo = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.btn-link.btn-condensed')))
            link_processo.click()
            time.sleep(3)
            log('Link do processo clicado com sucesso.')

            driver.switch_to.window(driver.window_handles[-1])
            time.sleep(3)
            driver.switch_to.default_content()
            campo = driver.find_element(By.ID, 'divTimeLine:txtPesquisa')
            campo.clear()
            campo.send_keys('Ato ordinatório praticado')
            time.sleep(2)
            botoes = driver.find_elements(By.TAG_NAME, 'button')
            for b in botoes:
                bid = b.get_attribute('id') or ''
                if 'btnPesquisa' in bid or 'pesquisa' in bid.lower():
                    b.click()
                    time.sleep(3)
                    break
            else:
                campo.send_keys(Keys.RETURN)
                time.sleep(3)
            ato = driver.find_element(By.XPATH, "//*[@id='divTimeLine:j_id403:0:j_id405:0:j_id416']/span")
            ato.click()
            time.sleep(5)
            log('Documento ATO aberto.')

            driver.switch_to.window(driver.window_handles[-1])
            time.sleep(2)
            driver.switch_to.default_content()
            frame = driver.find_element(By.ID, 'frameHtml')
            driver.switch_to.frame(frame)
            time.sleep(2)
            body = driver.find_element(By.TAG_NAME, 'body')
            texto = body.text
            driver.switch_to.default_content()

            dados_extraidos = extrair_dados_documento(texto)
            dados.update(dados_extraidos)
            if not dados.get('nr_processo'):
                dados['nr_processo'] = nr_processo
            log(f'Dados extraidos: {dados}')

            if len(driver.window_handles) >= 2:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
        except Exception as e:
            log(f'ERRO no processo {nr_processo}: {e}')
            if len(driver.window_handles) >= 2:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])

        resultados.append(dados)

    salvar_resultado(resultados, CAMINHO_SAIDA)
    log(f'Finalizado. Total: {len(resultados)}')
    com_dados = sum(1 for r in resultados if r.get('profissao'))
    log(f'Com dados: {com_dados} | Sem dados: {len(resultados) - com_dados}')
    print('\nNavegador permanece aberto. Pressione ENTER no terminal para fechar.')
    input()
    driver.quit()


if __name__ == '__main__':
    main()
