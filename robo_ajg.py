import os, re, time, subprocess, unicodedata
from datetime import datetime
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException, StaleElementReferenceException,
    UnexpectedAlertPresentException, NoAlertPresentException,
)
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
import pyperclip

CAMINHO_SAIDA = r'C:\Users\pi100348\Documents\scrapping_pje\resultado.xlsx'


def log(msg):
    print(f'[{datetime.now().strftime("%H:%M:%S")}] {msg}')


def remover_acentos(texto):
    return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')


def esperar_ajg(driver):
    try:
        el = driver.find_element(By.ID, 'loadingDiv')
        return el.value_of_css_property('display') == 'none'
    except (NoSuchElementException, StaleElementReferenceException):
        return True


def esperar_ajg_com_timeout(driver, timeout=120):
    WebDriverWait(driver, timeout).until(esperar_ajg)


def main():
    options = Options()
    options.add_argument('--start-maximized')
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(10)

    driver.get('https://ajg1.cjf.jus.br/aj/seguranca/efetuarloginintranet/efetuarLoginIntranet_efetuarLogin.jsf')
    time.sleep(3)
    log('Pagina do AJG aberta. Faca login manualmente.')
    input('Pressione ENTER apos estar logado no AJG... ')
    log('Login AJG confirmado. Ambos os sistemas prontos.')

    driver.switch_to.window(driver.window_handles[0])
    time.sleep(2)

    try:
        el = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.ID, 'link_tela_com_tooltip_1nomeacaodeprofissionais'))
        )
        print(f'Elemento encontrado: {el.tag_name} - {el.text[:50]}')
        el.click()
        print('Clicado com sucesso!')
    except TimeoutException:
        print('Elemento NAO encontrado. Verificando o que existe na pagina...')
        links = driver.find_elements(By.TAG_NAME, 'a')
        for l in links:
            lid = l.get_attribute('id') or ''
            ltext = l.text.strip()[:50]
            if 'nomeacao' in lid.lower() or 'nomeacao' in ltext.lower() or 'profission' in lid.lower():
                print(f'  id={lid} texto={ltext}')

    # Ler dados do resultado.xlsx
    wb = openpyxl.load_workbook(CAMINHO_SAIDA, read_only=True, data_only=True)
    ws = wb.active
    cabecalhos = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    resultados_excel = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dados_row = {}
        for col_idx, coluna in enumerate(cabecalhos):
            dados_row[coluna] = str(row[col_idx]) if row[col_idx] else ''
        if dados_row.get('nr_processo'):
            resultados_excel.append(dados_row)
    wb.close()

    log(f'Processos lidos do resultado.xlsx: {len(resultados_excel)}')

    for i, dados in enumerate(resultados_excel, 1):
        log(f'=== Processo {i}/{len(resultados_excel)}: {dados.get("nr_processo")} ===')

        try:
            # Clicar em "Novo" (se nao estiver na lista, voltar para ela antes)
            try:
                driver.find_element(By.ID, 'formAJIntranet:novo')
            except NoSuchElementException:
                log('Voltando para a pagina de nomeacoes...')
                driver.find_element(By.ID, 'link_tela_com_tooltip_1nomeacaodeprofissionais').click()
                esperar_ajg_com_timeout(driver)
                time.sleep(1)

            driver.find_element(By.ID, 'formAJIntranet:novo').click()
            esperar_ajg_com_timeout(driver)
            time.sleep(1)

            Select(WebDriverWait(driver, 60).until(
                EC.visibility_of_element_located((By.ID, 'formAJIntranet:id_categoriasProfissionais'))
            )).select_by_visible_text('PERITO')

            profissoes = Select(WebDriverWait(driver, 60).until(
                EC.visibility_of_element_located((By.ID, 'formAJIntranet:id_profissoes'))
            ))
            for option in profissoes.options:
                if remover_acentos(option.text).lower() == remover_acentos(dados['profissao']).lower():
                    profissoes.select_by_visible_text(option.text)
                    break
            else:
                raise Exception(f"Profissao '{dados['profissao']}' nao encontrada no seletor do AJG.")
            log(f'  OK: profissao={dados["profissao"]}')

            driver.find_element(By.ID, 'formAJIntranet:id_dataNomeacao').send_keys(dados['data_nomeacao'])
            log(f'  OK: data_nomeacao={dados["data_nomeacao"]}')

            if 'INSS' in dados['polo_passivo']:
                driver.find_element(By.XPATH,
                    '/html/body/form/div[3]/span/table/tbody/tr[7]/td[2]/fieldset/span/span/input[1]').click()
            else:
                driver.find_element(By.XPATH,
                    '/html/body/form/div[3]/span/table/tbody/tr[7]/td[2]/fieldset/span/span/input[2]').click()
            log(f'  OK: marquei polo_passivo={dados["polo_passivo"]}')

            driver.find_element(By.ID, 'formAJIntranet:concBenefAssistDeficBenefPrevIncLab').click()
            log('  OK: marquei concBenef')
            log('  Dados da nomeacao preenchidos.')

            driver.find_element(By.ID, 'formAJIntranet:avancar').click()

            try:
                WebDriverWait(driver, 120).until(esperar_ajg)
            except UnexpectedAlertPresentException:
                log('Este processo ja teve nomeacao paga pelo Poder Executivo. Pulando...')
                driver.find_element(By.ID, 'formAJIntranet:cancelar').click()
                try:
                    driver.switch_to.alert.accept()
                except NoAlertPresentException:
                    pass

            campo = WebDriverWait(driver, 60).until(
                EC.visibility_of_element_located((By.ID, 'formAJIntranet:id_NumeroProcessoJudicial'))
            )
            campo.click()
            time.sleep(0.5)

            # 1. Testar se o pyperclip + Ctrl+V funciona
            pyperclip.copy(dados['nr_processo'])
            print(f'Clipboard contem: {repr(dados["nr_processo"])}')
            ActionChains(driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
            time.sleep(2)

            # 2. Ler o valor do campo para ver se foi colado
            valor_campo = campo.get_attribute('value')
            print(f'Valor no campo apos colar: {repr(valor_campo)}')

            # 3. Mandar TAB
            campo.send_keys(Keys.TAB)
            time.sleep(3)
            print(f'URL apos TAB: {driver.current_url}')

            # 4. Verificar se apareceu alerta
            try:
                WebDriverWait(driver, 120).until(esperar_ajg)
            except TimeoutException:
                try:
                    WebDriverWait(driver, 30).until(esperar_ajg)
                except TimeoutException:
                    pass
                driver.refresh()
                try:
                    alerta = driver.switch_to.alert
                    alerta.accept()
                except NoAlertPresentException:
                    pass
                tentativa += 1
                continue
            except UnexpectedAlertPresentException:

                log('Este processo judicial ja teve nomeacao paga pelo Poder Executivo. Vou passar para o proximo')
            except TimeoutException:
                print('Timeout no loading')
                driver.find_element(By.ID, 'formAJIntranet:cancelar').click()
                try:
                    alerta = driver.switch_to.alert
                    alerta.accept()
                except NoAlertPresentException:
                    pass

            # Voltar para a pagina inicial (inicio do sistema) para o proximo processo
            try:
                driver.find_element(By.ID, 'formAJIntranet:bt_inicio_sistema').click()
                esperar_ajg_com_timeout(driver)
                time.sleep(1)
            except Exception:
                pass

        except Exception as e:
            log(f'  ERRO: {e}')
            try:
                driver.find_element(By.ID, 'formAJIntranet:bt_inicio_sistema').click()
                esperar_ajg_com_timeout(driver)
                time.sleep(1)
            except Exception:
                pass

    log('Finalizado.')
    if len(resultados_excel) == 1:
        log('Apenas 1 processo processado. O navegador permanece aberto para conferencia manual.')
    else:
        log(f'{len(resultados_excel)} processos processados. Todos concluidos.')
    print('\nPressione ENTER para fechar o navegador.')
    input()
    driver.quit()


if __name__ == '__main__':
    main()
